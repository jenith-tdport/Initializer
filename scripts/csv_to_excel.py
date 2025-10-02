#!/usr/bin/env python3
import csv
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Read CSV file
csv_file = '/home/user/webapp/public/data/initializer.csv'
excel_file = '/home/user/webapp/public/data/initializer_questions.xlsx'

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Quiz Questions"

# Read CSV and write to Excel
with open(csv_file, 'r', encoding='utf-8') as f:
    csv_reader = csv.reader(f)
    for row_idx, row in enumerate(csv_reader, 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

# Format the header row
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col in range(1, 11):  # 10 columns
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Adjust column widths
column_widths = {
    'A': 10,  # term
    'B': 12,  # kind
    'C': 12,  # category
    'D': 50,  # question
    'E': 35,  # optionA
    'F': 35,  # optionB
    'G': 35,  # optionC
    'H': 35,  # optionD
    'I': 8,   # correct
    'J': 60   # explanation
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Add alternating row colors for better readability
light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
for row in range(2, ws.max_row + 1):
    if row % 2 == 0:
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = light_fill

# Freeze the header row
ws.freeze_panes = 'A2'

# Add data validation for 'correct' column (only A, B, C, D allowed)
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"A,B,C,D"', allow_blank=False)
dv.add(f'I2:I{ws.max_row}')
ws.add_data_validation(dv)

# Save the Excel file
wb.save(excel_file)
print(f"Excel file created: {excel_file}")

# Also create a Google Sheets compatible version with instructions
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Instructions"

instructions = [
    ["How to Use This Spreadsheet with the Initializer Game"],
    [""],
    ["Option 1: Use with Google Sheets (Recommended)"],
    ["1. Upload this file to Google Sheets"],
    ["2. Go to File > Share > Publish to web"],
    ["3. Choose 'Quiz Questions' sheet and 'Comma-separated values (.csv)' format"],
    ["4. Copy the published URL"],
    ["5. Use this URL in the game's sheetUrl parameter"],
    [""],
    ["Option 2: Use as Static File"],
    ["1. Edit questions in Excel or Google Sheets"],
    ["2. Save/Export as CSV format"],
    ["3. Upload to your web server at /data/initializer.csv"],
    [""],
    ["CSV Format Requirements:"],
    ["- Headers must be: term,kind,category,question,optionA,optionB,optionC,optionD,correct,explanation"],
    ["- 'correct' column must contain only: A, B, C, or D"],
    ["- 'kind' can be: initialism, acronym, or backronym"],
    ["- No empty rows between questions"],
    [""],
    ["Tips for Good Questions:"],
    ["- Make all options plausible to increase difficulty"],
    ["- Keep explanations concise but educational"],
    ["- Balance correct answers across A, B, C, D options"],
    ["- Test your questions in the game after updating"]
]

for row_idx, row_data in enumerate(instructions, 1):
    for col_idx, value in enumerate(row_data, 1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

# Format instructions
title_font = Font(size=14, bold=True)
section_font = Font(size=12, bold=True)
ws2.cell(row=1, column=1).font = title_font
ws2.cell(row=3, column=1).font = section_font
ws2.cell(row=10, column=1).font = section_font
ws2.cell(row=15, column=1).font = section_font
ws2.cell(row=21, column=1).font = section_font

ws2.column_dimensions['A'].width = 100

# Add the questions sheet
ws_questions = wb2.create_sheet("Quiz Questions")

# Copy data from the first workbook
with open(csv_file, 'r', encoding='utf-8') as f:
    csv_reader = csv.reader(f)
    for row_idx, row in enumerate(csv_reader, 1):
        for col_idx, value in enumerate(row, 1):
            ws_questions.cell(row=row_idx, column=col_idx, value=value)

# Apply the same formatting
for col in range(1, 11):
    cell = ws_questions.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for col, width in column_widths.items():
    ws_questions.column_dimensions[col].width = width

for row in range(2, ws_questions.max_row + 1):
    if row % 2 == 0:
        for col in range(1, 11):
            ws_questions.cell(row=row, column=col).fill = light_fill

ws_questions.freeze_panes = 'A2'

# Save the complete Excel file
complete_file = '/home/user/webapp/initializer_questions_complete.xlsx'
wb2.save(complete_file)
print(f"Complete Excel file with instructions created: {complete_file}")